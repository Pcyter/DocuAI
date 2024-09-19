import argparse
import os

import chardet
import pythoncom
import tiktoken
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from win32com.client import Dispatch
from langchain_core.documents import Document as Doc

from myLogger import logger

# 定义大模型的token限制（例如GPT-4的token限制）
TOKEN_LIMIT = 6000  # 假设模型最大支持6000个tokens

def get_token_count(text, model_name="gpt-3.5-turbo"):
    """
    使用 tiktoken 计算输入文本的 token 数量。
    :param text: 输入文本
    :param model_name: 模型名称，默认为 gpt-3.5-turbo
    :return: 输入文本的 token 数量
    """
    encoding = tiktoken.encoding_for_model(model_name)
    tokens = encoding.encode(text)
    return len(tokens)

def split_text_into_chunks(text, chunk_size, model_name="gpt-3.5-turbo"):
    """
    将长文本分块，每块不超过指定的 token 长度。
    :param text: 输入文本
    :param chunk_size: 每个块的最大 token 数量
    :param model_name: 模型名称，默认为 gpt-3.5-turbo
    :return: 分块后的文本列表
    """
    encoding = tiktoken.encoding_for_model(model_name)
    tokens = encoding.encode(text)

    # 分块处理
    chunks = []
    for i in range(0, len(tokens), chunk_size):
        chunk_tokens = tokens[i:i + chunk_size]
        chunk_text = encoding.decode(chunk_tokens)
        chunks.append(chunk_text)

    return chunks

# 1. 解析不同类型的文档
def parse_docx(file_path):
    doc = Document(file_path)
    full_text = [paragraph.text for paragraph in doc.paragraphs]
    return "\n".join(full_text)


def parse_doc(file_path):
    pythoncom.CoInitialize()
    word = Dispatch("Word.Application")
    word.Visible = False
    doc = word.Documents.Open(file_path)
    doc_text = doc.Content.Text
    doc.Close()
    word.Quit()
    return doc_text


def parse_xlsx(file_path):
    wb = load_workbook(file_path)
    full_text = []
    for sheet in wb.sheetnames:
        sheet = wb[sheet]
        for row in sheet.iter_rows(values_only=True):
            full_text.append(" ".join([str(cell) for cell in row if cell is not None]))
    return "\n".join(full_text)


def parse_xls(file_path):
    pythoncom.CoInitialize()
    excel = Dispatch("Excel.Application")
    excel.Visible = False
    wb = excel.Workbooks.Open(file_path)
    full_text = []
    for sheet in wb.Sheets:
        for row in sheet.UsedRange.Rows:
            row_values = [str(cell.Value) for cell in row]
            full_text.append(" ".join(row_values))
    wb.Close()
    excel.Quit()
    return "\n".join(full_text)


def parse_pdf(file_path):
    reader = PdfReader(file_path)
    full_text = [page.extract_text() for page in reader.pages]
    return "\n".join(full_text)

def parse_txt(file_path):
    # 首先检测文件编码
    with open(file_path, 'rb') as f:
        raw_data = f.read()
        result = chardet.detect(raw_data)
        encoding = result['encoding']

    with open(file_path, "r", encoding=encoding) as file:
        full_text = file.read()
        return "".join(full_text)

def parse_file(file_path):
    if file_path.endswith(".docx"):
        return parse_docx(file_path)
    elif file_path.endswith(".doc"):
        return parse_doc(file_path)
    elif file_path.endswith(".xlsx"):
        return parse_xlsx(file_path)
    elif file_path.endswith(".xls"):
        return parse_xls(file_path)
    elif file_path.endswith(".pdf"):
        return parse_pdf(file_path)
    elif file_path.endswith(".txt"):
        return parse_txt(file_path)
    else:
        logger.error(f"Unsupported file format: {file_path}")
        raise ValueError(f"Unsupported file format: {file_path}")


def walkFile(file):
    file_list = []
    for root, dirs, files in os.walk(file):
        # root 表示当前正在访问的文件夹路径
        # dirs 表示该文件夹下的子目录名list
        # files 表示该文件夹下的文件list
        for f in files:
            file_list.append(os.path.join(root, f))
    return file_list


##文件是否存在判别函数
def file_exists(path):
    if not os.path.isfile(path):
        logger.error("文件 %s 不存在!" % path)
        raise argparse.ArgumentTypeError("文件 %s 不存在!" % path)
    return path


##目录是否存在判别函数
def path_exists(path):
    if not os.path.isdir(path):
        logger.error("目录 %s 不存在!" % path)
        raise argparse.ArgumentTypeError("目录 %s 不存在!" % path)
    return path

def parse_file_documents(file_path):
    try:
        text_all = parse_file(file_exists(file_path))
        logger.info("parse_file(file_exists(file_path)) end!")
        documents = []
        # 对每个块进行模型推理处理
        token_count = get_token_count(text_all)
        logger.info("token_count: %d" % token_count)
        if token_count < TOKEN_LIMIT:
            documents.append(Doc(page_content=f"{text_all}", metadata={"title": f"{file_path}"}))
        else:
            chunks = split_text_into_chunks(text_all, TOKEN_LIMIT)
            for check in chunks:
                documents.append(Doc(page_content=f"{check}", metadata={"title": f"{file_path}"}))

        logger.info("documents: %d" % len(documents))
        return documents
    except Exception as e:
        logger.error(f"Error is {e}")
